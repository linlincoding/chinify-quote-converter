#!/usr/bin/env python3
"""
Chinify Quote Converter — Internal Web App
Upload any client PDF or Excel quote → download Chinify-format Excel
"""

import os, re, io, json, sys, tempfile
from pathlib import Path

import streamlit as st

# Ensure the app's directory is on the path so convert_quote can be imported
sys.path.insert(0, str(Path(__file__).parent))

from convert_quote import (
    claude_extract_all_pages,
    claude_extract_from_image,
    annotate_yranges,
    extract_and_match_images,
    score_products,
    write_excel,
    MODEL,
)

import anthropic
import openpyxl

# Fixed paths (relative to this file)
_DIR = Path(__file__).parent
TEMPLATE_PATH = str(_DIR / "2025 J0 初始报价表.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# Excel extraction — handles arbitrary client Excel templates via Claude
# ══════════════════════════════════════════════════════════════════════════════
EXCEL_EXTRACT_PROMPT = """You are a furniture order data extractor.

I will give you the raw cell contents of a client's furniture quote spreadsheet.
Extract ALL product/furniture rows and return them as a JSON array.

Each product object must have these fields:
  "area"      — room or zone name (e.g. "SALA", "COMEDOR", "TERRAZA"). Use "" if not shown.
  "product"   — furniture name (e.g. "SOFÁ 3 PLAZAS", "MESA COMEDOR 8 PERS")
  "code"      — product code / SKU (e.g. "SL-2023"). Use "" if none.
  "qty"       — quantity as a string (e.g. "2", "1 SET"). Use "" if none.
  "dims"      — dimensions (e.g. "2400×900×750"). Use "" if none.
  "material"  — fabric / material code (e.g. "LP9086F-1"). Use "" if none.
  "extra_materials" — list of additional material codes for the same product (if any). Default [].

Rules:
- Skip column headers, grand totals, shipping rows, payment terms, company addresses.
- Area / room labels are NOT products — capture them in the "area" field for subsequent products.
- If a product spans multiple rows (e.g. two fabric options), keep the first row as the
  main product and put additional material codes in "extra_materials".
- Normalize dimensions to digits with × separator when possible.

Return ONLY a valid JSON array. No markdown, no explanation.

Spreadsheet contents:
{data}
"""


def _read_excel_as_text(excel_path: str) -> str:
    """Flatten all sheets into a plain-text representation for Claude."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if any(v for v in vals):          # skip fully blank rows
                rows.append(" | ".join(vals))
        if rows:
            sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(sections)


def extract_from_excel(excel_path: str) -> list[dict]:
    """Use Claude to extract product list from any Excel template."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        st.error("⚠ ANTHROPIC_API_KEY is not set — cannot extract from Excel.")
        return []

    text = _read_excel_as_text(excel_path)
    client = anthropic.Anthropic(api_key=api_key)

    with client.messages.stream(
        model=MODEL,
        max_tokens=4000,
        thinking={"type": "adaptive"},
        messages=[{
            "role": "user",
            "content": EXCEL_EXTRACT_PROMPT.format(data=text),
        }]
    ) as stream:
        resp = stream.get_final_message()

    raw = next(b.text for b in resp.content if b.type == "text").strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    products = json.loads(raw)

    # Normalise to match the shape that write_excel expects
    for p in products:
        p.setdefault("area", "")
        p.setdefault("product", "")
        p.setdefault("code", "")
        p.setdefault("qty", "")
        p.setdefault("dims", "")
        p.setdefault("material", "")
        p.setdefault("extra_materials", [])
        p.setdefault("y0", 0)
        p.setdefault("y1", 0)
        p.setdefault("swatches", [])
        p.setdefault("photo", None)
        p.setdefault("page", 1)
        p["source_type"] = "excel"

    return products


# ══════════════════════════════════════════════════════════════════════════════
# Core conversion — PDF or Excel → product list → Chinify Excel bytes
# ══════════════════════════════════════════════════════════════════════════════
def convert_file(uploaded_file, use_claude_clean: bool, status) -> tuple[bytes | None, list, dict]:
    """
    Process an uploaded Streamlit file object.
    Returns (excel_bytes, products, summary) or (None, [], {}) on failure.
    """
    suffix = Path(uploaded_file.name).suffix.lower()

    with tempfile.TemporaryDirectory() as tmpdir:
        in_path  = os.path.join(tmpdir, uploaded_file.name)
        out_path = os.path.join(tmpdir, "chinify_output.xlsx")

        with open(in_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # ── Extract products ────────────────────────────────────────────────
        if suffix == ".pdf":
            status.write("🤖 Claude Vision: reading all pages...")
            products = claude_extract_all_pages(in_path)
            status.write(f"✓ Found **{len(products)}** products across all pages")

            status.write("📐 Locating row positions for image matching...")
            products = annotate_yranges(in_path, products)

            status.write("🖼 Extracting & matching images...")
            products = extract_and_match_images(in_path, products)
            n_photo  = sum(1 for p in products if p.get("photo"))
            n_swatch = sum(1 for p in products if p.get("swatches"))
            status.write(f"✓ Photos: {n_photo}  |  Swatches: {n_swatch}")

        elif suffix in (".xlsx", ".xls"):
            status.write("📊 Sending Excel to Claude for extraction...")
            products = extract_from_excel(in_path)
            if not products:
                return None, [], {}
            status.write(f"✓ Found **{len(products)}** products")

        elif suffix in (".png", ".jpg", ".jpeg"):
            status.write("🖼 Reading screenshot/photo with Claude Vision...")
            products = claude_extract_from_image(in_path)
            if not products:
                return None, [], {}
            status.write(f"✓ Found **{len(products)}** products")

        else:
            st.error(f"Unsupported file type: {suffix}")
            return None, [], {}

        status.write("🧠 Scoring rows for manual review...")
        products = score_products(products)
        review_count = sum(1 for p in products if p.get("needs_review"))
        avg_confidence = (
            sum(p.get("confidence", 0.0) for p in products) / len(products)
            if products else 0.0
        )
        summary = {
            "product_count": len(products),
            "review_count": review_count,
            "avg_confidence": avg_confidence,
        }
        status.write(
            f"✓ Review queue: {review_count} row(s)  |  "
            f"Average confidence: {avg_confidence:.0%}"
        )

        # ── Write Chinify Excel ──────────────────────────────────────────────
        status.write("📝 Writing Chinify Excel format...")
        write_excel(products, TEMPLATE_PATH, out_path)
        status.write("✓ Excel generated")

        with open(out_path, "rb") as f:
            return f.read(), products, summary


# ══════════════════════════════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="Chinify Quote Converter",
        page_icon="🏠",
        layout="centered",
    )

    # ── Branding ─────────────────────────────────────────────────────────────
    st.title("🏠 Chinify Quote Converter")
    st.markdown(
        "Upload a client quote (PDF or Excel) and download the converted "
        "Chinify format Excel in seconds."
    )
    st.divider()

    # ── API key (env var preferred; fallback to input) ────────────────────────
    if not os.environ.get("ANTHROPIC_API_KEY"):
        key = st.text_input(
            "Anthropic API Key",
            type="password",
            placeholder="sk-ant-...",
            help="Set the ANTHROPIC_API_KEY environment variable to skip this field.",
        )
        if key:
            os.environ["ANTHROPIC_API_KEY"] = key
        else:
            st.info("Enter your API key to enable AI-powered extraction.")

    # ── Upload & options ──────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Upload client quote",
        type=["pdf", "xlsx", "xls", "png", "jpg", "jpeg"],
        help="PDF, Excel, or a screenshot/photo of the quote.",
        label_visibility="collapsed",
    )

    # Claude Vision is always used for PDF extraction (no longer optional)
    use_claude = True

    if not uploaded:
        st.markdown(
            """
            **Supported inputs:**
            - 📄 PDF — client quote PDFs
            - 📊 Excel (`.xlsx` / `.xls`) — any client template format
            - 🖼 Image (`.png` / `.jpg` / `.jpeg`) — screenshots or photos

            **Recommended workflow:**
            - Let AI prefill the Chinify template
            - Review only the rows marked as risky
            - Treat this as a time-saver, not a zero-check autopilot
            """
        )
        return

    # File info
    size_kb = uploaded.size / 1024
    st.markdown(f"**File:** `{uploaded.name}` — {size_kb:.1f} KB")

    if st.button("🔄 Convert", type="primary", use_container_width=True):
        output_bytes = None
        products = []
        summary = {}

        with st.status("Converting…", expanded=True) as status:
            try:
                output_bytes, products, summary = convert_file(uploaded, use_claude, status)
                if output_bytes:
                    status.update(label="Conversion complete ✅", state="complete", expanded=False)
                else:
                    status.update(label="Conversion failed ❌", state="error")
            except Exception as e:
                status.update(label=f"Error: {e}", state="error")
                st.exception(e)

        if output_bytes and products:
            col1, col2, col3 = st.columns(3)
            col1.metric("Products", summary.get("product_count", len(products)))
            col2.metric("Needs review", summary.get("review_count", 0))
            col3.metric("Avg confidence", f"{summary.get('avg_confidence', 0):.0%}")

            review_rows = [p for p in products if p.get("needs_review")]
            if review_rows:
                st.warning(
                    f"{len(review_rows)} row(s) should be checked before sending to factory."
                )

            # ── Product preview table ─────────────────────────────────────
            with st.expander("📋 Preview extracted products", expanded=True):
                rows = []
                for p in products:
                    rows.append({
                        "Product":  p.get("product", ""),
                        "Code":     p.get("code",    ""),
                        "Qty":      p.get("qty",     ""),
                        "Dims":     p.get("dims",    ""),
                        "Material": p.get("material",""),
                        "Confidence": f"{p.get('confidence', 0):.0%}",
                        "Review":   "Yes" if p.get("needs_review") else "",
                        "Photo":    "✓" if p.get("photo") else "",
                        "Swatch":   f"{len(p.get('swatches',[]))}x" if p.get("swatches") else "",
                    })
                st.dataframe(rows, use_container_width=True, hide_index=True)

            if review_rows:
                with st.expander("🛠 Rows that need manual review", expanded=True):
                    review_table = []
                    for p in review_rows:
                        review_table.append({
                            "Product": p.get("product", ""),
                            "Code": p.get("code", ""),
                            "Qty": p.get("qty", ""),
                            "Dims": p.get("dims", ""),
                            "Material": p.get("material", ""),
                            "Confidence": f"{p.get('confidence', 0):.0%}",
                            "Why review": " | ".join(p.get("review_reasons", [])),
                        })
                    st.dataframe(review_table, use_container_width=True, hide_index=True)

            # ── Download ──────────────────────────────────────────────────
            out_name = Path(uploaded.name).stem + "_chinify.xlsx"
            st.download_button(
                label="⬇️ Download Chinify Excel",
                data=output_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

    # ── Footer ────────────────────────────────────────────────────────────────
    st.divider()
    st.caption("Chinify Internal Tool — v1.0")


if __name__ == "__main__":
    main()
