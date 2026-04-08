#!/usr/bin/env python3
"""
Chinify Quote Converter — Claude Vision Primary Extraction
- Claude Vision extracts ALL product data from ANY PDF format, ANY number of pages
- pdfplumber used only for y-position lookup (image matching), NOT text parsing
- Supports multi-page PDFs, any column layout, any language
"""

import os, re, io, json, base64, shutil, zipfile, unicodedata
from datetime import date

import fitz
import pdfplumber
import anthropic
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ── Config ────────────────────────────────────────────────────────────────────
PDF_PATH      = "COTIZACION VIDALTA - PEDIDOS CHINA(1).pdf"
TEMPLATE_PATH = "2025 J0 初始报价表.xlsx"
OUTPUT_PATH   = "VIDALTA_chinify_报价.xlsx"
MODEL         = "claude-opus-4-6"

# ── Column layout — matches J078 EXACTLY ─────────────────────────────────────
# B=Número  C=Área  D=Producto  E=Foto  F=Largo  G=Prof  H=Altura
# I=Detalle  J=Cantidad  K=PrecioUnit  L=PrecioTotal  M=CBM  N=KG  O=TotalCBM  P=Muestra
COL_WIDTHS = {
    "A": 1.5, "B": 14, "C": 13, "D": 24, "E": 20,
    "F": 9,   "G": 10, "H": 9,  "I": 22, "J": 8,
    "K": 12,  "L": 12, "M": 7,  "N": 7,  "O": 9,  "P": 18,
}
FOTO_W_PX   = 168
FOTO_H_PX   = 126
NOTAS_W_PX  = 150
NOTAS_H_PX  = 126
ROW_H_IMG   = 100
ROW_H_TEXT  = 25

# ── Styles ────────────────────────────────────────────────────────────────────
def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
AREA_FILL   = PatternFill("solid", fgColor="D6DCE4")
FILL_W      = PatternFill("solid", fgColor="FFFFFF")
FILL_B      = PatternFill("solid", fgColor="EBF3FB")
TOTAL_FILL  = PatternFill("solid", fgColor="D9E2F3")
TERMS_FILL  = PatternFill("solid", fgColor="F2F7FC")

WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
DARK_BOLD  = Font(name="Calibri", bold=True, color="1F3864", size=9)
DARK_NORM  = Font(name="Calibri", color="1F3864", size=9)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

STANDARD_FIELDS = {
    "area",
    "code",
    "product",
    "qty",
    "dims",
    "detail",
    "material",
    "page",
    "extra_materials",
    "extra_fields",
    "swatches",
    "photo",
    "y0",
    "y1",
    "source_type",
    "confidence",
    "review_reasons",
    "needs_review",
    "_swatch",
    "_extra",
}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1  Claude Vision — primary extraction from ALL pages (any PDF format)
# ══════════════════════════════════════════════════════════════════════════════
EXTRACT_PROMPT = """You are a furniture BOQ (Bill of Quantities) extractor.

I'm showing you all pages of a furniture specification document.
Extract EVERY furniture item and return them as a JSON array.

Each item must have these exact fields:
  "code"     — product code/SKU (e.g. "CH-03", "DS-01", "SL-2023"). Use "" if none.
  "product"  — product name (e.g. "VISITOR CHAIR", "SOFÁ 3 PLAZAS"). Fix any OCR errors.
  "qty"      — quantity as a string (e.g. "4", "2", "1 SET"). Use "" if none.
  "area"     — room/zone name (e.g. "MINISTER OFFICE", "SALA", "TERRAZA"). Use "" if not shown.
  "dims"     — full dimension string exactly as written (e.g. "3000W x 400D x 780H CM",
               "OVERALL: 79W x 76D x 80H SEAT HEIGHT: 46CM", "45DIA X55H"). Use "" if none.
  "detail"   — visible description/spec/detail text for this product. Use "" if none.
  "material" — all material notes for this product joined with " | "
               (e.g. "LEATHER: TO BE SELECT | WOOD: TO BE SELECT"). Use "" if none.
  "extra_fields" — object of any other visible client-specific fields that do not fit the schema.
  "page"     — page number (1-indexed integer) where this item appears.

Rules:
- Extract ALL items across ALL pages — do not skip any.
- If the same CODE appears in multiple rooms with different quantities, create a SEPARATE entry for each.
- Section/area headers (e.g. "MINISTER OFFICE", "MINISTER BEDROOM", "SALA") are NOT products —
  use them to fill "area" for the products listed below them.
- Skip: column header rows, page footers, company notes, legal text.
- "dims": copy the full dimension text exactly as written — do not reformat or shorten.
- "material": join ALL material lines for that product row with " | ".
- Put any additional client-only fields into "extra_fields" using the visible label as the key.

Return ONLY a valid JSON array. No markdown, no explanation.
"""

IMAGE_EXTRACT_PROMPT = """You are a furniture quote extractor.

I am showing you a client quote as one screenshot or photo.
Extract every furniture/product line you can see and return them as a JSON array.

Each item must have these exact fields:
  "code"     — product code/SKU. Use "" if none.
  "product"  — product name. Use "" if none.
  "qty"      — quantity as a string. Use "" if none.
  "area"     — room/zone name. Use "" if not shown.
  "dims"     — dimensions exactly as written. Use "" if none.
  "detail"   — visible description/spec/detail text for this product. Use "" if none.
  "material" — all material/fabric/finish notes for the same item joined with " | ". Use "" if none.
  "extra_fields" — object of any other visible client-specific fields that do not fit the schema.
  "page"     — use 1.

Rules:
- Extract only real product rows/items, not totals, headers, addresses, or payment terms.
- If one product spans multiple lines, merge the lines into one item when possible.
- If text is blurry or uncertain, still return the best guess, but do not invent values that are not visible.
- Put any additional client-only fields into "extra_fields" using the visible label as the key.

Return ONLY a valid JSON array. No markdown, no explanation.
"""


BATCH_SIZE = 3   # pages per Claude call — keeps token usage manageable


def _parse_json_response(raw: str) -> list[dict]:
    raw = raw.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


def _canonical_key(label: str) -> str:
    text = unicodedata.normalize("NFKD", label or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9]+", "", text).lower()
    return text


def _key_to_standard_field(label: str) -> str | None:
    key = _canonical_key(label)
    alias_map = {
        "clave": "code",
        "codigo": "code",
        "code": "code",
        "sku": "code",
        "ref": "code",
        "reference": "code",
        "nombre": "product",
        "name": "product",
        "product": "product",
        "producto": "product",
        "detalle": "detail",
        "detail": "detail",
        "descripcion": "detail",
        "description": "detail",
        "spec": "detail",
        "specification": "detail",
        "cantidad": "qty",
        "cant": "qty",
        "canttotal": "qty",
        "qty": "qty",
        "quantity": "qty",
        "area": "area",
        "ubicacion": "area",
        "location": "area",
        "colocacion": "area",
        "medida": "dims",
        "medidas": "dims",
        "tamano": "dims",
        "tamaño": "dims",
        "dimension": "dims",
        "dimensions": "dims",
        "dim": "dims",
        "material": "material",
        "fabric": "material",
        "tela": "material",
    }
    return alias_map.get(key)


def _harmonize_product_fields(product: dict) -> None:
    cleaned_extra = {}
    extra_fields = product.get("extra_fields") or {}
    if not isinstance(extra_fields, dict):
        extra_fields = {}

    candidate_items = list(extra_fields.items())
    for key, val in list(product.items()):
        if key in STANDARD_FIELDS:
            continue
        candidate_items.append((key, val))

    for raw_key, raw_val in candidate_items:
        value = "" if raw_val is None else str(raw_val).strip()
        if not value:
            continue

        std_key = _key_to_standard_field(str(raw_key))
        if std_key:
            current = str(product.get(std_key, "") or "").strip()
            if not current:
                product[std_key] = value
                continue
            if current == value:
                continue

        cleaned_extra[str(raw_key).strip()] = value

    product["extra_fields"] = cleaned_extra


def _normalize_products(products: list[dict], source_type: str) -> list[dict]:
    normalized = []
    for p in products:
        p.setdefault("area", "")
        p.setdefault("code", "")
        p.setdefault("product", "")
        p.setdefault("qty", "")
        p.setdefault("dims", "")
        p.setdefault("detail", "")
        p.setdefault("material", "")
        p.setdefault("page", 1)
        p.setdefault("extra_materials", [])
        p.setdefault("extra_fields", {})
        p.setdefault("swatches", [])
        p.setdefault("photo", None)
        p.setdefault("y0", 0)
        p.setdefault("y1", 0)
        p["source_type"] = source_type
        _harmonize_product_fields(p)
        normalized.append(p)
    return normalized


def _call_claude_batch(client, page_images: list[tuple[int, bytes]]) -> list[dict]:
    """
    Send a batch of (page_num, png_bytes) to Claude and return extracted products.
    page_num is 1-indexed (absolute page number in the full document).
    """
    content = []
    for page_num, png_bytes in page_images:
        png_b64 = base64.standard_b64encode(png_bytes).decode()
        content.append({"type": "text", "text": f"=== PAGE {page_num} ==="})
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": png_b64},
        })
    content.append({"type": "text", "text": EXTRACT_PROMPT})

    resp = client.messages.create(
        model=MODEL,
        max_tokens=8000,
        messages=[{"role": "user", "content": content}],
    )

    text_blocks = [b.text for b in resp.content if b.type == "text"]
    if not text_blocks:
        raise RuntimeError(
            f"Claude returned no text (batch pages {[p for p,_ in page_images]}). "
            f"Stop reason: {resp.stop_reason}"
        )

    return _parse_json_response(text_blocks[0])


def claude_extract_all_pages(pdf_path: str) -> list[dict]:
    """
    Render ALL PDF pages and extract products via Claude Vision.
    Processes in batches of BATCH_SIZE pages to stay within token limits.
    Works with any PDF format, any number of pages, any language.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY not set")

    # Render all pages at 96 DPI — sufficient for text, lower token usage
    doc = fitz.open(pdf_path)
    mat = fitz.Matrix(96 / 72, 96 / 72)
    page_images = []
    for i in range(len(doc)):
        pix = doc[i].get_pixmap(matrix=mat, alpha=False)
        page_images.append((i + 1, pix.tobytes("png")))   # 1-indexed page num
    doc.close()

    client = anthropic.Anthropic(api_key=api_key)
    all_products: list[dict] = []

    for batch_start in range(0, len(page_images), BATCH_SIZE):
        batch = page_images[batch_start: batch_start + BATCH_SIZE]
        page_nums = [p for p, _ in batch]
        print(f"  Extracting pages {page_nums}...")
        products = _call_claude_batch(client, batch)
        all_products.extend(products)

    return _normalize_products(all_products, source_type="pdf")


def claude_extract_from_image(image_path: str) -> list[dict]:
    """
    Extract products from a quote screenshot or photo.
    This is less reliable than native PDF/Excel parsing, so downstream review is important.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY not set")

    suffix = os.path.splitext(image_path)[1].lower()
    media_type = "image/png" if suffix == ".png" else "image/jpeg"
    with open(image_path, "rb") as f:
        raw_bytes = f.read()
    img_b64 = base64.standard_b64encode(raw_bytes).decode()

    client = anthropic.Anthropic(api_key=api_key)
    resp = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": img_b64,
                    },
                },
                {"type": "text", "text": IMAGE_EXTRACT_PROMPT},
            ],
        }],
    )

    text_blocks = [b.text for b in resp.content if b.type == "text"]
    if not text_blocks:
        raise RuntimeError(f"Claude returned no text. Stop reason: {resp.stop_reason}")
    products = _normalize_products(_parse_json_response(text_blocks[0]), source_type="image")
    return _attach_screenshot_row_photos(products, raw_bytes)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2  Annotate y-ranges via pdfplumber (for image matching ONLY)
# Matches by product CODE — no hard-coded column positions
# ══════════════════════════════════════════════════════════════════════════════
_CODE_RE = re.compile(r'^([A-Z]{2,6}-\d+[A-Z]?(?:\.[AB])?)', re.IGNORECASE)


def annotate_yranges(pdf_path: str, products: list[dict]) -> list[dict]:
    """
    Find each product's y0/y1 on its page by matching its CODE in pdfplumber tables.
    Products whose code can't be located keep y0=y1=0 (no image extraction attempted).
    """
    code_y: dict[tuple, tuple] = {}  # {(page_0idx, code_upper): (y0, y1)}

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.find_tables()
            if not tables:
                continue
            t = tables[0]
            for row in t.rows:
                if not row.cells:
                    continue
                for cell_bbox in row.cells[:4]:   # CODE is usually in first 4 cols
                    if cell_bbox is None:
                        continue
                    try:
                        text = (page.crop(cell_bbox).extract_text() or "").strip()
                    except Exception:
                        continue
                    m = _CODE_RE.match(text)
                    if m:
                        code = m.group(1).upper()
                        y0 = round(cell_bbox[1], 2)
                        y1 = round(cell_bbox[3], 2)
                        key = (page_num, code)
                        if key not in code_y:
                            code_y[key] = (y0, y1)
                        break

    for p in products:
        page_idx = p.get("page", 1) - 1
        code = p.get("code", "").upper()
        if code:
            key = (page_idx, code)
            if key in code_y:
                p["y0"], p["y1"] = code_y[key]

    matched = sum(1 for p in products if p["y0"] != 0 or p["y1"] != 0)
    print(f"  Y-range matched: {matched}/{len(products)} products")
    return products


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3  Extract images and match to products by page + y-overlap
# No hard-coded x-coordinates — scans the full page width
# ══════════════════════════════════════════════════════════════════════════════
def _is_blank(pix, threshold=247):
    samples = pix.samples
    return (sum(samples) / len(samples)) > threshold


def _crop_looks_blank(pil_img: PILImage.Image) -> bool:
    gray = pil_img.convert("L")
    extrema = gray.getextrema()
    if not extrema:
        return True
    low, high = extrema
    return (high - low) < 12


def _to_jpeg_bytes(pil_img: PILImage.Image) -> bytes:
    if pil_img.mode != "RGB":
        pil_img = pil_img.convert("RGB")
    buf = io.BytesIO()
    pil_img.save(buf, format="JPEG", quality=88)
    return buf.getvalue()


def _attach_screenshot_row_photos(products: list[dict], raw_bytes: bytes) -> list[dict]:
    """
    Heuristic fallback for screenshot inputs that contain one product image per row.
    Crops the likely image column and splits it by product order.
    """
    if not products:
        return products
    if any(p.get("photo") for p in products):
        return products

    pil = PILImage.open(io.BytesIO(raw_bytes))
    if pil.mode not in ("RGB", "RGBA"):
        pil = pil.convert("RGB")

    width, height = pil.size
    count = len(products)
    if count == 1:
        x0 = int(width * 0.62)
        x1 = int(width * 0.81)
        y0 = int(height * 0.18)
        y1 = int(height * 0.90)
        crop = pil.crop((x0, y0, x1, y1))
        if not _crop_looks_blank(crop):
            products[0]["photo"] = _to_jpeg_bytes(crop)
        return products

    # Screenshot quotes usually place the product picture in a dedicated image column.
    # Keep the crop focused on that column and preserve most of each row height.
    x0 = int(width * 0.62)
    x1 = int(width * 0.81)
    content_top = int(height * 0.17)
    content_bottom = int(height * 0.86)
    row_h = max(1, (content_bottom - content_top) / count)

    for idx, product in enumerate(products):
        y0 = int(content_top + idx * row_h + row_h * 0.02)
        y1 = int(content_top + (idx + 1) * row_h - row_h * 0.02)
        crop = pil.crop((x0, y0, x1, y1))
        if _crop_looks_blank(crop):
            continue
        product["photo"] = _to_jpeg_bytes(crop)

    return products


def extract_and_match_images(pdf_path: str, products: list[dict]) -> list[dict]:
    """
    Multi-page image extraction:
    - Pass 1: embedded images matched by y-overlap (no x-coordinate assumptions)
    - Pass 2: render-based fallback for the right ~35% of the page (picture column area)
    LEFT-column images → photo (furniture photo, shown in FOTO/E column)
    RIGHT-column images → swatches (fabric, shown in MUESTRA/P column)
    For PDFs with a single centre PICTURE column, all go to photo.
    """
    doc = fitz.open(pdf_path)

    by_page: dict[int, list[dict]] = {}
    for p in products:
        pg = p.get("page", 1) - 1
        by_page.setdefault(pg, []).append(p)

    mat3 = fitz.Matrix(3, 3)

    for pg in range(len(doc)):
        if pg not in by_page:
            continue
        page = doc[pg]
        page_w = page.rect.width
        page_prods = [p for p in by_page[pg] if p["y0"] != 0 or p["y1"] != 0]
        if not page_prods:
            continue

        # ── Pass 1: embedded images ──────────────────────────────────────────
        emb = []
        for img in page.get_images(full=True):
            xref = img[0]
            try:
                bbox = page.get_image_bbox(img)
                base_img = doc.extract_image(xref)
            except Exception:
                continue
            if len(base_img["image"]) < 8_000:
                continue
            xc = (bbox.x0 + bbox.x1) / 2
            yc = (bbox.y0 + bbox.y1) / 2
            emb.append({
                "xc": xc, "yc": yc,
                "data": base_img["image"],
                "size": len(base_img["image"]),
            })

        for p in page_prods:
            y0, y1 = p["y0"], p["y1"]
            hits = [i for i in emb if y0 <= i["yc"] <= y1]
            if not hits:
                continue

            # Split hits into left (< 45% page width) and right (>= 45%)
            left_hits  = [i for i in hits if i["xc"] < page_w * 0.45]
            right_hits = [i for i in hits if i["xc"] >= page_w * 0.45]

            if left_hits and right_hits:
                # Two-column layout (e.g. VIDALTA): left=photo, right=swatches
                p["photo"] = max(left_hits, key=lambda x: x["size"])["data"]
                for hit in sorted(right_hits, key=lambda x: x["yc"]):
                    p["swatches"].append(hit["data"])
            elif left_hits:
                p["photo"] = max(left_hits, key=lambda x: x["size"])["data"]
            elif right_hits:
                # Single PICTURE column (BOQ style): centre/right image → photo only
                p["photo"] = max(right_hits, key=lambda x: x["size"])["data"]

        # ── Pass 2: render fallback for rows with no embedded image ──────────
        for p in page_prods:
            if p.get("photo"):
                continue
            y0, y1 = p["y0"], p["y1"]
            if y0 == 0 and y1 == 0:
                continue
            # Try right half of page first (picture column), then left
            for x0_frac, x1_frac in [(0.45, 0.78), (0.05, 0.43)]:
                pix = page.get_pixmap(
                    matrix=mat3,
                    clip=fitz.Rect(page_w * x0_frac, y0, page_w * x1_frac, y1),
                    alpha=False,
                )
                if not _is_blank(pix):
                    p["photo"] = pix.tobytes("jpeg", jpg_quality=85)
                    break

    doc.close()

    prods_with_photo  = sum(1 for p in products if p.get("photo"))
    prods_with_swatch = sum(1 for p in products if p.get("swatches"))
    print(f"  Furniture photos : {prods_with_photo}/{len(products)}")
    print(f"  Fabric swatches  : {prods_with_swatch} products "
          f"({sum(len(p['swatches']) for p in products)} total)")
    return products


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4  Parse dimensions
# Handles labelled W/D/H, diameter formats, and plain NxNxN
# ══════════════════════════════════════════════════════════════════════════════
def parse_dims(s: str) -> tuple[str, str, str]:
    """
    Extract (width/L, depth/D, height/H) from a dimension string.
    Examples handled:
      "3000W x 400D x 780H CM"           → ("3000", "400", "780")
      "OVERALL: 79W x 76D x 80H ..."     → ("79", "76", "80")
      "45DIA X55H"                        → ("Ø45", "Ø45", "55")
      "Ø cm 60 x h 50"                   → ("Ø60", "Ø60", "50")
      "2400×900×750"                      → ("2400", "900", "750")
    """
    if not s:
        return "", "", ""

    # Explicit W / D / H labels
    w = re.search(r'(\d+(?:\.\d+)?)\s*W\b', s, re.IGNORECASE)
    d = re.search(r'(\d+(?:\.\d+)?)\s*D\b', s, re.IGNORECASE)
    h = re.search(r'(\d+(?:\.\d+)?)\s*H\b', s, re.IGNORECASE)
    if w and h:
        return w.group(1), (d.group(1) if d else ""), h.group(1)

    # Diameter format: "45DIA", "Ø60", "ø cm 60"
    dia = re.search(r'(?:DIA|Ø|ø)\s*(?:cm\s*)?(\d+(?:\.\d+)?)', s, re.IGNORECASE)
    h2  = re.search(r'\bh\s*(\d+(?:\.\d+)?)', s, re.IGNORECASE)
    if dia:
        v = f"Ø{dia.group(1)}"
        return v, v, (h2.group(1) if h2 else "")

    # Generic N×N×N / N*N*N / N x N x N
    parts = re.split(r'[×*]|\s+[xX]\s+', s.strip().rstrip('*'))
    parts = [
        re.sub(r'\s*(CM|MM|M)\b.*$', '', p, flags=re.IGNORECASE).strip()
        for p in parts if p.strip() and re.search(r'\d', p)
    ]
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return parts[0], parts[1], ""

    return s[:30] if s else "", "", ""


def score_products(products: list[dict]) -> list[dict]:
    """
    Add lightweight review metadata so the UI can focus humans on risky rows.
    This is heuristic, not model-native confidence.
    """
    for p in products:
        reasons = []
        score = 0.2

        product = (p.get("product") or "").strip()
        code = (p.get("code") or "").strip()
        qty = (p.get("qty") or "").strip()
        dims = (p.get("dims") or "").strip()
        material = (p.get("material") or "").strip()
        source_type = p.get("source_type", "")

        if product:
            score += 0.35
        else:
            reasons.append("Missing product name")

        if qty:
            score += 0.15
        else:
            reasons.append("Missing quantity")

        if code:
            score += 0.12
        else:
            reasons.append("Missing product code")

        if dims:
            score += 0.10
            l, w, h = parse_dims(dims)
            if not l:
                reasons.append("Dimensions could not be split cleanly")
            elif not (w or h):
                reasons.append("Dimensions look incomplete")
        else:
            reasons.append("Missing dimensions")

        if material:
            score += 0.08
        else:
            reasons.append("Missing material / fabric")

        if p.get("photo"):
            score += 0.08
        if p.get("swatches"):
            score += 0.05
        if source_type == "image":
            score -= 0.10
            reasons.append("Screenshot/photo input is OCR-sensitive")
        elif source_type == "excel":
            score += 0.05

        product_len = len(product)
        if 0 < product_len < 4:
            reasons.append("Product name is very short")
            score -= 0.08

        if not code and not dims and not material:
            reasons.append("Too little structured detail to trust auto-fill")
            score -= 0.12

        if not reasons:
            reasons.append("Looks consistent")

        confidence = max(0.0, min(score, 0.99))
        p["confidence"] = round(confidence, 2)
        p["review_reasons"] = reasons
        p["needs_review"] = confidence < 0.75 or any(
            reason.startswith("Missing") for reason in reasons
        )
    return products


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5  Write Excel — J078 format, fabric swatch in MUESTRA/P, photo in FOTO/E
# ══════════════════════════════════════════════════════════════════════════════
def make_xl_img(raw_bytes, target_w, target_h):
    """Fit image into exact target_w x target_h canvas (white bg, centered)."""
    pil = PILImage.open(io.BytesIO(raw_bytes))
    if pil.mode in ("RGBA", "P", "LA"):
        pil = pil.convert("RGB")
    pil.thumbnail((target_w - 4, target_h - 4), PILImage.LANCZOS)
    canvas = PILImage.new("RGB", (target_w, target_h), (255, 255, 255))
    ox = (target_w - pil.width) // 2
    oy = (target_h - pil.height) // 2
    canvas.paste(pil, (ox, oy))
    buf = io.BytesIO()
    canvas.save(buf, format="JPEG", quality=85, dpi=(96, 96))
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = target_w
    xl.height = target_h
    return xl


def _write_cell(ws, col, row, val, fill, font, alignment, border):
    c = ws.cell(row=row, column=col if isinstance(col, int) else
                openpyxl.utils.column_index_from_string(col))
    c.value = val; c.fill = fill; c.font = font
    c.alignment = alignment; c.border = border


def _collect_extra_columns(products: list[dict]) -> list[str]:
    ordered = []
    seen = set()

    def add_key(key: str):
        label = (key or "").strip()
        if not label:
            return
        if label in seen:
            return
        seen.add(label)
        ordered.append(label)

    for p in products:
        extra_fields = p.get("extra_fields") or {}
        if isinstance(extra_fields, dict):
            for key in extra_fields.keys():
                add_key(str(key))

        for key in p.keys():
            if key not in STANDARD_FIELDS:
                add_key(str(key))

    preferred = {
        "clave2": 0,
        "model": 0,
        "proveedor": 1,
        "provider": 1,
        "supplier": 1,
        "remark": 2,
        "remarks": 2,
        "nota": 2,
        "notes": 2,
    }
    ordered.sort(key=lambda label: (preferred.get(_canonical_key(label), 99), label.lower()))
    return ordered


def _build_detail_text(product: dict) -> str:
    detail = (product.get("detail") or "").strip()
    material = (product.get("material") or "").strip()
    if detail and material:
        return f"{detail} | {material}"
    return detail or material


def write_excel(products, template_path, out_path):
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["家具清单"]

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws["H5"] = date.today()

    # Unmerge rows 9+
    to_unmerge = [str(mr) for mr in list(ws.merged_cells.ranges) if mr.min_row >= 9]
    for mr in to_unmerge:
        ws.unmerge_cells(mr)

    extra_columns = _collect_extra_columns(products)
    extra_start_idx = openpyxl.utils.column_index_from_string("Q")
    extra_col_map = {
        key: get_column_letter(extra_start_idx + idx)
        for idx, key in enumerate(extra_columns)
    }

    for col in extra_col_map.values():
        ws.column_dimensions[col].width = 18

    # Reset ALL cells rows 9-300 to plain white
    _blank = PatternFill("solid", fgColor="FFFFFF")
    _no_border = Border()
    for r in range(9, 301):
        ws.row_dimensions[r].height = ROW_H_TEXT
        for c in range(1, extra_start_idx + len(extra_columns)):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill  = _blank
            cell.border = _no_border
            cell.font  = DARK_NORM

    # Headers rows 9-10
    ws.row_dimensions[9].height  = 26
    ws.row_dimensions[10].height = 20
    for r in (9, 10):
        for c in range(1, extra_start_idx + len(extra_columns)):
            ws.cell(row=r, column=c).fill = HEADER_FILL

    for col, lbl in {
        "B": "NÚMERO",        "C": "ÁREA\n区域",
        "D": "PRODUCTO\n产品", "E": "FOTO\n照片",
        "I": "DETALLE\n细节",  "J": "CANTIDAD\n量",
        "K": "P.UNIT\nUSD",   "L": "P.TOTAL\nUSD",
        "M": "CBM",           "N": "KG",
        "O": "TOTAL\nCBM",    "P": "MUESTRA\n皮料",
    }.items():
        ws.merge_cells(f"{col}9:{col}10")
        c = ws[f"{col}9"]
        c.value = lbl; c.font = WHITE_BOLD; c.fill = HEADER_FILL
        c.alignment = CENTER; c.border = bdr()

    ws.merge_cells("F9:H9")
    ws["F9"].value = "TAMAÑO / 尺寸 (MM)"
    ws["F9"].font = WHITE_BOLD; ws["F9"].fill = HEADER_FILL
    ws["F9"].alignment = CENTER; ws["F9"].border = bdr()
    for col, lbl in [("F","LARGO"),("G","PROFUNDIDAD"),("H","ALTURA")]:
        c = ws[f"{col}10"]
        c.value = lbl; c.font = WHITE_BOLD; c.fill = HEADER_FILL
        c.alignment = CENTER; c.border = bdr()

    for key, col in extra_col_map.items():
        ws.merge_cells(f"{col}9:{col}10")
        c = ws[f"{col}9"]
        c.value = str(key).upper()
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = bdr()

    # Expand products with multiple swatches into separate rows
    rows_data = []
    for p in products:
        swatches   = p.get("swatches", [])
        extra_mats = p.get("extra_materials", [])

        if len(swatches) <= 1:
            rows_data.append({**p, "_swatch": swatches[0] if swatches else None, "_extra": False})
        else:
            rows_data.append({**p, "_swatch": swatches[0], "_extra": False})
            for i, sw in enumerate(swatches[1:]):
                mat_i = extra_mats[i] if i < len(extra_mats) else ""
                rows_data.append({**p, "_swatch": sw, "_extra": True,
                                   "code": "", "dims": "", "qty": "",
                                   "material": mat_i})

    # Product rows
    row       = 11
    alt       = 0

    for p in rows_data:
        fill = FILL_W if alt % 2 == 0 else FILL_B
        has_photo  = bool(p.get("photo"))
        has_swatch = bool(p.get("_swatch"))
        ws.row_dimensions[row].height = ROW_H_IMG if (has_photo or has_swatch) else ROW_H_TEXT

        l, w, h = parse_dims(p.get("dims", ""))
        detail_text = _build_detail_text(p)

        data_cols = {
            "B": p.get("code",""),
            "C": "" if p.get("_extra") else p.get("area", ""),
            "D": p.get("product",""),
            "E": "",
            "F": l, "G": w, "H": h,
            "I": detail_text,
            "J": p.get("qty",""),
            "K": "", "L": "", "M": "", "N": "", "O": "",
            "P": "",
        }
        for col, val in data_cols.items():
            c = ws[f"{col}{row}"]
            c.value = val; c.fill = fill; c.font = DARK_NORM; c.border = bdr()
            c.alignment = LEFT if col in ("D","I","P") else CENTER

        extra_fields = p.get("extra_fields") or {}
        for key, col in extra_col_map.items():
            val = ""
            if isinstance(extra_fields, dict) and key in extra_fields:
                val = extra_fields.get(key, "")
            elif key in p and key not in STANDARD_FIELDS:
                val = p.get(key, "")

            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = fill
            c.font = DARK_NORM
            c.border = bdr()
            c.alignment = LEFT

        # Furniture photo → FOTO/E
        if has_photo and not p.get("_extra"):
            try:
                xl = make_xl_img(p["photo"], FOTO_W_PX, FOTO_H_PX)
                xl.anchor = f"E{row}"
                ws.add_image(xl)
            except Exception as e:
                print(f"  ⚠ Photo [{p['product'][:20]}]: {e}")

        # Fabric swatch → MUESTRA/P
        if has_swatch:
            try:
                xl = make_xl_img(p["_swatch"], NOTAS_W_PX, NOTAS_H_PX)
                xl.anchor = f"P{row}"
                ws.add_image(xl)
            except Exception as e:
                print(f"  ⚠ Swatch [{p['product'][:20]}]: {e}")

        row += 1; alt += 1

    # TOTAL section
    for lbl in ["TOTAL", "SHIPPING", "TOTAL GENERAL"]:
        ws.row_dimensions[row].height = 22
        total_end_col = get_column_letter(max(openpyxl.utils.column_index_from_string("J"), extra_start_idx + len(extra_columns) - 1))
        ws.merge_cells(f"B{row}:{total_end_col}{row}")
        c = ws[f"B{row}"]
        c.value = lbl; c.font = DARK_BOLD; c.fill = TOTAL_FILL
        c.alignment = CENTER; c.border = bdr()
        for cidx in range(openpyxl.utils.column_index_from_string("K"), extra_start_idx + len(extra_columns)):
            col = get_column_letter(cidx)
            ws[f"{col}{row}"].fill = TOTAL_FILL
            ws[f"{col}{row}"].border = bdr()
        row += 1

    # Terms
    ws.row_dimensions[row].height = 22
    terms_end_col = get_column_letter(max(openpyxl.utils.column_index_from_string("P"), extra_start_idx + len(extra_columns) - 1))
    ws.merge_cells(f"B{row}:{terms_end_col}{row}")
    ws[f"B{row}"].value = "TÉRMINOS / 条款"
    ws[f"B{row}"].font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    ws[f"B{row}"].fill  = HEADER_FILL; ws[f"B{row}"].alignment = LEFT
    row += 1

    for term in [
        "1. El Precio es EXW, puesto en la fábrica.",
        "2. Todos los precios se basan en especificaciones. Nos reservamos el derecho de cambiar el precio según los requisitos adicionales del cliente.",
        "3. El tiempo de producción se cuenta a partir del día que se confirme todos los detalles.",
        "4. El 30/50% debe pagarse mediante T/T como depósito. El restante después de inspección.",
        "5. Los muebles deben estar terminados en 45/60 días después de recibir el depósito.",
        "6. Empaquetado estándar de fábrica. Empaquetados especiales serán coste extra.",
        "7. Los productos son personalizados y no se devolverán salvo defectos de fabricación.",
    ]:
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"B{row}:{terms_end_col}{row}")
        c = ws[f"B{row}"]
        c.value = term; c.font = Font(name="Calibri", size=8, color="555555")
        c.fill = TERMS_FILL; c.alignment = LEFT
        row += 1

    ws.freeze_panes = None
    ws.sheet_view.view = "normal"
    wb.save(out_path)

    _fix_view_xml(out_path, last_data_row=row)
    print(f"\n✓  Saved: {out_path}")


def _fix_view_xml(path, last_data_row):
    """Patch worksheet XML to force normal view."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                xml = data.decode("utf-8")
                xml = re.sub(r'view="pageBreakPreview"', 'view="normal"', xml)
                xml = re.sub(r'\s*zoomScaleNormal="[^"]*"', '', xml)
                xml = re.sub(r'\s*zoomScalePageLayoutView="[^"]*"', '', xml)
                xml = re.sub(r'zoomScale="[^"]*"', 'zoomScale="100"', xml)
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN (CLI usage)
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Chinify Quote Converter — Claude Vision extraction")
    print("=" * 60)

    print("\n[1/3] Claude Vision: extracting products from all pages...")
    products = claude_extract_all_pages(PDF_PATH)
    print(f"  Products found: {len(products)}")

    print("\n[2/3] Annotating y-ranges for image matching...")
    products = annotate_yranges(PDF_PATH, products)

    print("\n[3/3] Extracting & matching images...")
    products = extract_and_match_images(PDF_PATH, products)

    print("\n  Final product list:")
    for p in products:
        sw = "🎨" if p.get("swatches") else "  "
        ph = "📷" if p.get("photo") else "  "
        print(f"  {ph}{sw} [p{p.get('page',1)}] [{p['area'][:18]:18}] "
              f"{p['product'][:28]:28} x{p['qty']:<5} {p['dims'][:20]:20}  {p['material'][:30]}")

    print(f"\n[4/3] Writing Excel...")
    write_excel(products, TEMPLATE_PATH, OUTPUT_PATH)

    print("\n" + "=" * 60)
    print("  SELF-AUDIT")
    print("=" * 60)
    print(f"\n  Total products         : {len(products)}")
    print(f"  With photo (FOTO/E)    : {sum(1 for p in products if p.get('photo'))}")
    print(f"  With swatch (MUESTRA/P): {sum(1 for p in products if p.get('swatches'))}")
    print(f"  No image at all        : {sum(1 for p in products if not p.get('photo') and not p.get('swatches'))}")
    print()
    print(f"  {'PRODUCT':<28} {'CODE':<14} {'QTY':<6} {'DIMS':<22} FOTO SWATCH")
    print(f"  {'-'*28} {'-'*14} {'-'*6} {'-'*22} ---- ------")
    for p in products:
        l, w, h = parse_dims(p.get("dims", ""))
        dims_str = f"{l}×{w}×{h}" if w else (l or "-")
        sw = f"{len(p.get('swatches', []))}x" if p.get("swatches") else "---"
        ph = "✓" if p.get("photo") else "---"
        print(f"  {p['product']:<28} {p.get('code',''):<14} {p.get('qty',''):<6} "
              f"{dims_str:<22} {ph:<4} {sw}")
    print("Done!")


if __name__ == "__main__":
    main()
